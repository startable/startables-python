import io
from pathlib import Path
from textwrap import dedent

import numpy as np
import openpyxl
import pandas as pd
import pytest
from pytest import fixture, raises
try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    # openpyxl < 2.6
    from openpyxl.worksheet import Worksheet
from openpyxl import load_workbook

from pyscheme import make_root_environment, Environment
from startables import ColumnMetadata
from startables.startables import Table, read_csv, read_excel, Bundle, nan
from startables.units import Unit, CustomUnitPolicy, ScaleUnitConversion


@fixture(scope='module')
def input_dir() -> Path:
    return Path(__file__).parent / 'input'


class TestTable:

    @fixture
    def some_df(self):
        return pd.DataFrame(data=[[nan, 'gnu', 3], [4, 'gnat', '{{(+ x y)}}']], columns=['a', 'b', 'c'])

    @fixture
    def col_specs(self):
        return {n: ColumnMetadata(Unit(u)) for n, u in zip(['a', 'b', 'c'], ['-', 'text', 'm'])}

    @fixture
    def col_specs_with_format(self):
        return {'a': ColumnMetadata(Unit('-'), format_str='${:,.2f}'),
                'b': ColumnMetadata(Unit('text')),
                'c': ColumnMetadata(Unit('m'), format_str='.4e')}

    @fixture
    def some_df_with_digits(self):
        return pd.DataFrame(data=[[nan, 'gnu', 3.23412121],
                                  [4.12, 'gnat', 0.023],
                                  [0.4, 'galah', 42.01],
                                  [0.04334, 'gentoo', 43232],
                                  [4000.04334, 'gerbil', 43232.0987]],
                            columns=['a', 'b', 'c'])

    @fixture
    def some_table(self, some_df, col_specs):
        return Table(df=some_df, name='some_table', col_specs=col_specs, destinations=['success', 'glory'])

    @fixture
    def some_table_with_digits(self, some_df_with_digits, col_specs_with_format):
        return Table(df=some_df_with_digits, name='some_table_with_digits', col_specs=col_specs_with_format,
                     destinations=['success', 'glory'])

    def test_init_with_col_specs(self, some_df, col_specs):
        t = Table(df=some_df, name='adequate_table', col_specs=col_specs)
        assert t.name == 'adequate_table'
        assert t.destinations == ['all']
        assert t.df.iloc[1, 0] == 4
        assert t.col_names == ['a', 'b', 'c']
        assert t.col_units == ['-', 'text', 'm']
        assert t.col_specs == col_specs
        assert len(t) == 2

    def test_init_with_no_col_specs_at_all(self, some_df):
        t = Table(df=some_df, name='adequate_table')
        cs = t.col_specs
        assert list(cs.keys()) == ['a', 'b', 'c']
        assert [cs[col_name].unit for col_name in cs] == ['text', 'text', 'text']

    def test_init_errors(self, some_df, col_specs):
        with pytest.raises(ValueError):
            too_short_col_specs = {n: col_specs[n] for n in col_specs if n != 'b'}
            Table(df=some_df, name='adequate_table', col_specs=too_short_col_specs)
        with pytest.raises(ValueError):
            destination_with_illegal_space = ['ok', 'definitely not ok!', 'ok_again']
            Table(df=some_df, name='adequate_table', destinations=destination_with_illegal_space)
        with raises(ValueError):
            destinations_with_illegal_duplicates = ['ok', 'you_again', 'you_again']
            Table(df=some_df, name='adequate_table', destinations=destinations_with_illegal_duplicates)

    def test_df_setter(self, some_table: Table):
        df_with_subset_of_columns = pd.DataFrame(data=[[nan, 3], [4, 666]], columns=['a', 'c'])
        some_table.df = df_with_subset_of_columns
        assert some_table.col_names == ['a', 'c']
        assert some_table.col_units == ['-', 'm']

        df_with_new_unknown_column = pd.DataFrame(data=[[nan, 'gnu', 3], [4, 'gnat', '{{(+ x y)}}']],
                                                  columns=['some_unknown_column', 'b', 'c'])
        with pytest.raises(ValueError):
            some_table.df = df_with_new_unknown_column

    def test_copy(self, some_table: Table):
        t1 = some_table
        t2 = some_table.copy()
        assert t1.df is not t2.df  # checking it's actually a new copy
        assert t2.name == t1.name
        assert t2.destinations == t1.destinations
        assert t2.col_names == t1.col_names
        assert t2.col_units == t1.col_units
        pd.testing.assert_frame_equal(t1.df, t2.df)

    def test_to_csv(self, some_table: Table):
        out = io.StringIO()
        some_table.to_csv(out)
        assert out.getvalue() == dedent("""\
            **some_table;;
            success glory
            a;b;c
            -;text;m
            -;gnu;3
            4.0;gnat;{{(+ x y)}}
            
            """)

    def test_to_csv_with_format(self, some_table_with_digits: Table):
        out = io.StringIO()
        some_table_with_digits.to_csv(out)
        print(out.getvalue())
        assert out.getvalue() == dedent("""\
            **some_table_with_digits;;
            success glory
            a;b;c
            -;text;m
            -;gnu;3.2341e+00
            $4.12;gnat;2.3000e-02
            $0.40;galah;4.2010e+01
            $0.04;gentoo;4.3232e+04
            $4,000.04;gerbil;4.3232e+04
            
            """)

    def test_to_csv_nonstring_colnames_and_destinations(self):
        # PS-53 Bundle.to_csv() fails when column names are not strings
        nonstring_colnames = [1.234, 666.0, 42.0]
        nonstring_destinations = [1984, 2001.2001]
        df = pd.DataFrame(data=[[nan, 'gnu', 3], [4, 'gnat', '{{(+ x y)}}']], columns=nonstring_colnames)
        col_specs = {n: ColumnMetadata(Unit(u)) for n, u in zip(nonstring_colnames, ['-', 'text', 'm'])}
        t = Table(df=df, name='some_table', col_specs=col_specs,
                  destinations=nonstring_destinations)
        out = io.StringIO()
        t.to_csv(out)
        assert out.getvalue() == dedent("""\
            **some_table;;
            1984 2001.2001
            1.234;666.0;42.0
            -;text;m
            -;gnu;3
            4.0;gnat;{{(+ x y)}}
            
            """)

    def test_to_excel(self, some_table: Table):
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        some_table.to_excel(ws)
        assert ws.cell(row=1, column=1).value == f'**{some_table.name}'
        assert ws.cell(2, 1).value == f'{" ".join(some_table.destinations)}'
        assert ws.cell(3, 2).value == 'b'
        assert ws.cell(4, 3).value == 'm'
        assert ws.cell(5, 1).value == '-'
        assert ws.cell(6, 3).value == '{{(+ x y)}}'

    def test_to_excel_with_digits(self, some_table_with_digits: Table):
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        some_table_with_digits.to_excel(ws)
        assert ws.cell(row=1, column=1).value == f'**{some_table_with_digits.name}'
        assert ws.cell(2, 1).value == f'{" ".join(some_table_with_digits.destinations)}'
        assert ws.cell(3, 2).value == 'b'
        assert ws.cell(4, 3).value == 'm'
        assert ws.cell(5, 1).value == '-'
        assert ws.cell(5, 3).value == '3.2341e+00'
        assert ws.cell(6, 1).value == '$4.12'
        assert ws.cell(6, 3).value == '2.3000e-02'
        assert ws.cell(7, 1).value == '$0.40'
        assert ws.cell(7, 3).value == '4.2010e+01'
        assert ws.cell(8, 1).value == '$0.04'
        assert ws.cell(8, 3).value == '4.3232e+04'
        assert ws.cell(9, 1).value == '$4,000.04'
        assert ws.cell(9, 3).value == '4.3232e+04'

    def test_evaluate_expressions(self, some_table: Table):
        env: Environment = make_root_environment().define('x', 42).define('y', 7)
        assert some_table.evaluate_expressions(env, inplace=False).df.iloc[1, 2] == 49

        context_dict = {'x': 7, 'y': 9}
        assert some_table.evaluate_expressions(context_dict, inplace=False).df.iloc[1, 2] == 16

        env.update([('y', 10)])
        some_table.evaluate_expressions(env, inplace=True)
        assert some_table.df.iloc[1, 2] == 52

    def test_evaluate_expression_syntax_error(self):
        df = pd.DataFrame(data=[[nan, 'gnu', 3], [4, 'gnat', '{{((((*+-/ x y heres_a_syntax_error!!!!!!!!!!!!!!!!}}']],
                          columns=['a', 'b', 'c'])
        col_specs = {n: ColumnMetadata(Unit(u)) for n, u in zip(['a', 'b', 'c'], ['-', 'text', 'm'])}
        t = Table(df=df, name='some_table', col_specs=col_specs, destinations=['success', 'glory'])
        with raises(SyntaxError, match=r"Syntax error in expression in table 'some_table', column 2, row 1"):
            t.evaluate_expressions({'x': 7, 'y': 9}, inplace=False)

    def test_convert_to_ref_units(self):
        df = pd.DataFrame([
            [11, 12, 13],
            [21, 22, 23]], columns=['a', 'b', 'c'])
        cs = {n: ColumnMetadata(Unit(u)) for n, u in zip(['a', 'b', 'c'], ['m', 'mm', 'km'])}
        t = Table(df, name='Fool', col_specs=cs)
        cup = CustomUnitPolicy([
            ScaleUnitConversion(Unit('mm'), Unit('m'), 0.001),
            ScaleUnitConversion(Unit('km'), Unit('m'), 1000)])
        t_ref = t.convert_to_ref_units(cup, inplace=False)
        assert t_ref.col_units == [Unit('m')] * 3
        assert (np.array(t_ref.df) == np.array([[11, 0.012, 13000],
                                                [21, 0.022, 23000]])).all()

    def test_convert_to_ref_units_unknown_unit(self):
        df = pd.DataFrame([
            [11, 12, 13],
            [21, 22, 23]], columns=['a', 'b', 'c'])
        cs = {n: ColumnMetadata(Unit(u)) for n, u in zip(['a', 'b', 'c'], ['m', 'mm', 'km'])}
        t = Table(df, name='Fool', col_specs=cs)
        cup_no_km = CustomUnitPolicy([ScaleUnitConversion(Unit('mm'), Unit('m'), 0.001)])
        t_ref_no_km = t.convert_to_ref_units(cup_no_km, inplace=False, units_not_in_policy='ignore')
        assert t_ref_no_km.col_units == [Unit('m'), Unit('m'), Unit('km')]
        assert (np.array(t_ref_no_km.df) == np.array([[11, 0.012, 13],
                                                      [21, 0.022, 23]])).all()

        with raises(ValueError):
            t.convert_to_ref_units(cup_no_km, inplace=False, units_not_in_policy='raise')

    def test_convert_to_home_units(self):
        df = pd.DataFrame([
            [11, 12, 13],
            [21, 22, 23]], columns=['a', 'b', 'c'])
        cs = {n: ColumnMetadata(Unit(u), Unit(hu)) for n, u, hu in zip(
            ['a', 'b', 'c'], ['m', 'm', 'm'], ['m', 'mm', 'km'])}
        t = Table(df, name='Fool', col_specs=cs)
        cup = CustomUnitPolicy([
            ScaleUnitConversion(Unit('mm'), Unit('m'), 0.001),
            ScaleUnitConversion(Unit('km'), Unit('m'), 1000)])
        t_home = t.convert_to_home_units(cup)
        assert t_home.col_units == [Unit('m'), Unit('mm'), Unit('km')]
        assert (np.array(t_home.df) == np.array([[11, 12000, 0.013],
                                                 [21, 22000, 0.023]])).all()

        # TODO tests for convert_units, and convert_to_home_units error handling of messed up unit policies


class TestBundle:

    @fixture
    def csv_path(self, input_dir: Path) -> Path:
        return input_dir / 'example.csv'

    @fixture
    def xlsx_path(self, input_dir: Path) -> Path:
        return input_dir / 'example.xlsx'

    @fixture
    def csv_path_illegal_empty_cell(self, input_dir: Path) -> Path:
        return input_dir / 'example_illegal_empty_cell.csv'

    @fixture
    def csv_path_illegal_str_in_num_col(self, input_dir: Path) -> Path:
        return input_dir / 'example_illegal_str_in_num_col.csv'

    @fixture
    def some_bundle(self, csv_path: Path):
        with open(str(csv_path)) as f:
            return read_csv(f)

    @fixture
    def csv_path_with_header(self, input_dir: Path) -> Path:
        return input_dir / 'example_header.csv'

    def test__filter_tables_noargs(self, some_bundle: Bundle):
        tables = some_bundle._filter_tables()
        assert len(tables) == 4

    @pytest.mark.parametrize("name,ignore_case,expected_len", [
        ('farm', True, 0),
        ('taxidermy', True, 1),
        ('TAXIdermy', True, 1),
        ('TAXIdermy', False, 0),
    ])
    def test__filter_tables_by_name(self, some_bundle: Bundle, name, ignore_case, expected_len):
        assert len(some_bundle._filter_tables(name=name, ignore_case=ignore_case)) == expected_len

    @pytest.mark.parametrize("name_pattern,ignore_case,expected_len", [
        ('farm', True, 1),
        ('fARM', True, 1),
        ('TAXIdermy', False, 0),
    ])
    def test__filter_tables_by_name_pattern(self, some_bundle: Bundle, name_pattern, ignore_case, expected_len):
        assert some_bundle._filter_tables(name_pattern='farm')[0].name == 'farm_animals'
        assert len(some_bundle._filter_tables(name_pattern=name_pattern, ignore_case=ignore_case)) == expected_len

    @pytest.mark.parametrize("destination,ignore_case,expected_len", [
        ('your_farm', True, 1),
        ('YOUR_farm', True, 1),
        ('YOUR_farm', False, 0),
    ])
    def test__filter_tables_by_destination(self, some_bundle: Bundle, destination, ignore_case, expected_len):
        assert some_bundle._filter_tables(destination='your_farm')[0].name == 'farm_animals'
        assert len(some_bundle._filter_tables(destination=destination, ignore_case=ignore_case)) == expected_len

    @pytest.mark.parametrize("destination_pattern,ignore_case,expected_len", [
        ('_farm', True, 1),
        ('_FaRm', True, 1),
        ('_FaRm', False, 0),
    ])
    def test__filter_tables_by_destination_pattern(self, some_bundle: Bundle, destination_pattern, ignore_case,
            expected_len):
        assert some_bundle._filter_tables(destination_pattern='_farm')[0].name == 'farm_animals'
        assert len(some_bundle._filter_tables(destination_pattern=destination_pattern,
                                              ignore_case=ignore_case)) == expected_len

    def test_filter(self, some_bundle: Bundle):
        assert len(some_bundle.filter().tables) == 4
        assert some_bundle.filter(destination='your_farm').tables[0].name == 'farm_animals'

    def test_pop_tables_noargs(self, csv_path: Path):
        with open(str(csv_path)) as f:
            b = read_csv(f)
        p = b.pop_tables()
        assert len(p) == 4
        assert len(b.tables) == 0

    def test_pop_tables_by_name(self, csv_path: Path):
        with open(str(csv_path)) as f:
            b = read_csv(f)
        taxidermy_tables = b.pop_tables(name='taxidermy')
        assert len(taxidermy_tables) == 1
        assert len(b.tables) == 3

    def test_copy(self, some_bundle: Bundle):
        b1 = some_bundle
        b2 = b1.copy()
        assert len(b1.tables) == len(b2.tables)
        assert b1.tables[1].name == b2.tables[1].name

    def test_read_csv(self, csv_path: Path):
        with open(str(csv_path)) as f:
            b = read_csv(f)
        assert len(b.tables) == 4
        t = b.tables[0]
        assert t.name == 'farm_animals'
        assert t.col_names == ['species', 'n_legs', 'avg_weight']
        assert t.col_units == ['text', '-', 'kg']
        assert t.destinations == ['your_farm', 'my_farm', 'farms_galore']
        df = t.df
        assert df.iloc[4, 2] == 9
        assert df.iloc[1, 2] == '{{(* age 30)}}'
        assert np.isnan(df.iloc[2, 2])
        assert np.isnan(df.iloc[3, 1])  # PS-15 Accept 'NaN' as NaN marker
        assert df.iloc[5, 0] == '1234'  # PS-28 Numerical data in text columns gets read
        assert df.shape == (6, 3)
        t2 = b.tables[2]
        assert t2.name == 'taxidermy'
        assert t2.df.iloc[3, 3] == pd.Timestamp('2012-05-01 12:34')
        assert pd.isna(t2.df.iloc[1, 3])
        assert t.evaluate_expressions({'age': 3}).df.iloc[1, 2] == 90

        assert len(b.tables[3]) == 0  # PS-5 Empty StarTables are unjustly ignored / omitted

    def test_read_csv_no_extra_delimiters_on_tables(self, input_dir):
        # PS-19 Reading from CSV can fail if not enough column delimiters on first line of CSV file
        with open(str(input_dir / 'example_comma_decimal.csv')) as f:
            b = read_csv(f)
        assert len(b.tables) == 3
        t = b.tables[0]
        df = t.df
        assert df.iloc[4, 2] == 9.876


    def test_read_csv_comma_decimal(self, input_dir):
        # PS-14  Can't control decimal separators when reading CSV files
        with open(str(input_dir / 'example_comma_decimal.csv')) as f:
            b = read_csv(f, decimal=',')
        assert len(b.tables) == 3
        t = b.tables[0]
        df = t.df
        assert df.iloc[4, 2] == 9.876

    def test_read_excel(self, xlsx_path: Path):
        # with open(str(xlsx_path)) as f:
        #     b = read_excel(f)  # <-- pd.read_excel() doesn't seem to work with file-like objects. Is doc wrong?
        b = read_excel(xlsx_path)
        assert len(b.tables) == 4
        t = b.tables[0]
        assert t.name == 'farm_animals'
        assert t.col_names == ['species', 'n_legs', 'avg_weight']
        assert t.col_units == ['text', '-', 'kg']
        assert t.destinations == ['your_farm', 'my_farm', 'farms_galore']
        df = t.df
        assert df.iloc[4, 2] == 9
        assert df.iloc[1, 2] == '{{(* age 30)}}'
        assert np.isnan(df.iloc[2, 2])
        assert np.isnan(df.iloc[3, 1])  # PS-15 Accept 'NaN' as NaN marker
        assert df.iloc[5, 0] == '1234'  # PS-28 Numerical data in text columns gets read
        assert df.shape == (6, 3)
        t2 = b.tables[2]  # Table on second sheet!
        assert t2.name == 'taxidermy'
        assert t2.df.iloc[3, 3] == pd.Timestamp('2012-05-01 12:34:56')
        assert pd.isna(t2.df.iloc[1, 3])
        assert t.evaluate_expressions({'age': 3}).df.iloc[1, 2] == 90

        assert len(b.tables[3]) == 0  # PS-5 Empty StarTables are ignored / omitted

    def test_read_illegal(self, input_dir: Path):
        with open(str(input_dir / 'example_illegal_empty_cell.csv')) as f:
            with pytest.raises(ValueError):
                read_csv(f)
        with open(str(input_dir / 'example_illegal_str_in_num_col.csv')) as f:
            with pytest.raises(ValueError):
                read_csv(f)

    def test_read_csv_missing_separators(self, input_dir: Path):
        with open(str(input_dir / 'example_missing_separators.csv')) as f:
            try:
                read_csv(f, missing_separators='ignore')  # 'ignore' is default
            except pd.errors.ParserError as e:
                pytest.fail(f'No exception should be thrown with missing_separators=\'ignore\'... {e}')

        # should fail with pandas.errors.ParserError
        with open(str(input_dir / 'example_missing_separators.csv')) as f:
            with pytest.raises(pd.errors.ParserError) as excinfo:
                read_csv(f, missing_separators='raise')
            assert "Error tokenizing data" in str(excinfo.value)

    def test_read_csv_utf8bom_with_open(self, input_dir: Path):
        # old approach where the user wraps with "with open"
        with open(str(input_dir / 'example_saveas_unicode_UTF8-BOM.csv')) as f:
            try:
                bundle = read_csv(f)
            except pd.errors.ParserError as e:
                pytest.fail(f'No exception should be thrown with missing_separators=\'ignore\'... {e}')

            assert len(bundle.tables) == 1  # load_ULS1_incl_LF

    def test_read_csv_utf8bom_without_open(self, input_dir: Path):
        # new approach where the user relys on startables/pandas to do the open/close
        bundle = read_csv(str(input_dir / 'example_saveas_unicode_UTF8-BOM.csv'))
        assert len(bundle.tables) == 1  # load_ULS1_incl_LF

    def test_read_csv_faulty(self, input_dir: Path):
        # new approach where the user relies on startables/pandas to do the open/close
        # should fail with ValueError
        with pytest.raises(ValueError) as excinfo:
            read_csv(str(input_dir / 'example_huge_file_missing_unit_and_values.csv'))
        assert "Illegal empty cell in numerical column '-' of table 'load_markovMatrixSetup" in str(excinfo.value)

    def test_read_csv_with_header(self, csv_path_with_header: Path):
        with open(str(csv_path_with_header)) as f:
            b = read_csv(f)
        assert len(b.tables) == 2
        t = b.tables[0]
        assert t.name == 'layer_cake'
        assert t.col_names == ['height', 'segment_top', 'diameter', 'colour']
        assert t.col_units == ['mm', 'mm', 'cm', 'text']
        assert t.destinations == ['birthday_party']
        df = t.df
        assert df.iloc[2, 0] == 65
        assert df.iloc[2, 1] == 75
        assert df.loc[0, 'colour'] == 'pink'

        t2 = b.tables[1]
        assert t2.name == 'beverages'
        assert t.destinations == ['birthday_party']
        df = t2.df
        assert df.loc[0, 'name'] == 'orange juice'
        assert df.loc[1, 'volume'] == 1000
        assert df.loc[3, 'quantity'] == 24


    # TODO test write datetime
    # TODO test read illegal empty cells, illegal str in num col, illegal values in datetime columns

    def test_to_csv(self, some_bundle: Bundle):
        out = io.StringIO()
        some_bundle.to_csv(out)
        assert out.getvalue() == dedent("""\
            **farm_animals;;;
            your_farm my_farm farms_galore
            species;n_legs;avg_weight
            text;-;kg
            chicken;2.0;3
            pig;4.0;{{(* age 30)}}
            goat;4.0;-
            cow;-;200
            goose;2.0;9
            1234;-;-
            
            **fruit;;;
            all
            kind;is_yummy
            text;onoff
            apple;0
            raspberry;1
            strawberry;1
            pineapple;0
            
            **taxidermy;;;
            all
            name;species;needs_repair;time_of_death
            text;text;onoff;datetime
            Sam;crow;1;2012-04-28 12:34:00
            Guy;mouse;0;-
            Kurt;ferret;0;2012-04-30 12:34:00
            Louise;rabbit;0;2012-05-01 12:34:00
            
            **empty_table;;;
            all
            foo;bar
            text;-
            
            """)

    def test_to_excel(self, some_bundle: Bundle, tmpdir):
        some_bundle.to_excel(tmpdir.join('some_bundle.xlsx'))

    def test_to_csv_with_header(self, some_bundle: Bundle):
        out = io.StringIO()
        some_bundle.to_csv(out, header='Info table\nWith:; farm animals, Fruit, etc.\t', sep=';')
        print(out.getvalue())
        assert out.getvalue() == dedent("""\
            Info table
            With:; farm animals, Fruit, etc.
            ;;;
            **farm_animals;;;
            your_farm my_farm farms_galore
            species;n_legs;avg_weight
            text;-;kg
            chicken;2.0;3
            pig;4.0;{{(* age 30)}}
            goat;4.0;-
            cow;-;200
            goose;2.0;9
            1234;-;-

            **fruit;;;
            all
            kind;is_yummy
            text;onoff
            apple;0
            raspberry;1
            strawberry;1
            pineapple;0

            **taxidermy;;;
            all
            name;species;needs_repair;time_of_death
            text;text;onoff;datetime
            Sam;crow;1;2012-04-28 12:34:00
            Guy;mouse;0;-
            Kurt;ferret;0;2012-04-30 12:34:00
            Louise;rabbit;0;2012-05-01 12:34:00

            **empty_table;;;
            all
            foo;bar
            text;-

            """)

    def test_to_excel_with_header(self, some_bundle: Bundle, tmpdir):

        header = ' Test Header\nDate:; Today\nNumeric Value:; 0.123\n'
        header_sep = ';'

        # write the bundle to excel
        some_bundle.to_excel(tmpdir.join('some_bundle.xlsx'), header=header, header_sep=header_sep)

        # now read in the excel table
        wb = load_workbook(filename=tmpdir.join('some_bundle.xlsx'), read_only=True)
        ws = wb.active
        # compare line/rows and their contents
        for line, row in zip(header.split('\n'), ws.rows):
            if line == '':
                for cell in row:
                    assert(cell.value == None)
            else:
                for col, cell in zip(line.split(header_sep), row):
                    assert (col == cell.value)

